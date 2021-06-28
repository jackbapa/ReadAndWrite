import pandas as pd
import openpyxl
import multiprocessing as m_p
import openpyxl.cell.cell as c




class Excel(m_p.Process):

    # 决定在写入时，是否保存的开关
    Save_in_Write = True
    # 默认的保存行，每多少行保存一次
    Save_each_rows = 5
    # 默认的工作sheet
    Work_on_sheet = "Sheet1"
    # 多sheet并发写入模式
    Write_many_sheets = False

    control =None

    def __init__(self, filepath_and_name):
        # super是把父类的方法当做自己的类的方法去调用
        # 具体的说，就相当于自己调用self.run()，就是从内部的调用，自己的类调用自己方法
        # 所以这里是不需要传入self到 __init__(self)中
        super().__init__()
        # EXCEL类继承了m_p.Process类，使用m_p.Process.__init__的方法是一个外部方法。
        # 外部方法是因为m_p.Process并不是己类的方法，就像在类内调用类外定义的一个函数
        # 调用外部的函数处理己类，需要传递EXCEL类进去，也就是要把self传入，如下
        # m_p.Process.__init__(self)
        self.reader_content = []
        self.queue_writer = m_p.Queue(100)
        self.control = m_p.Queue(1)
        # 为多sheet准备的接口
        self.queue_for_sheets = [m_p.Queue(3)]
        self.FirstRow = True
        self.count = 0
        self.ItemIndex = {}
        self.filepath_and_name = filepath_and_name
        try :
            self.book = openpyxl.load_workbook(filepath_and_name)


            self.sheets_names = self.book.sheetnames
            self.sheet_row_cloum = {str(temp): self.book[temp].max_row for temp in self.sheets_names}
        except Exception as e:
            wb = openpyxl.Workbook()

            wb.save(filepath_and_name)
            # print("#*"*10)
            # print(e)
            # print(filepath_and_name)
            self.book = openpyxl.load_workbook(filepath_and_name)
            self.sheets_names = self.book.sheetnames
            if "Sheet1" not in self.sheets_names:
                self.book.create_sheet("Sheet1")
                self.book.save(filepath_and_name)
            # print(self.sheets_names)
            self.sheet_row_cloum = {str(temp): self.book[temp].max_row for temp in self.sheets_names}
            self.sheets = {str(temp): self.book[temp] for temp in self.sheets_names}

    def write_row_Save(self, sheet_name, data_list_or_dic, save_row=5):
        # data_list_or_dic支持dic  --  list  --   panda.series
        # save_row 时隔多少行保存一次

        # print("data is   ",data_list_or_dic)
        # print("sheet is ",sheet_name)
        self.write_row(sheet_name, data_list_or_dic)
        if self.count % save_row == 0:
            # print('sacvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvve')
            self.book.save(self.filepath_and_name)

    def write_row(self, sheet_name, data_list_or_dic):
        # data_list_or_dic支持dic  --  list  --   panda.series
        # 需要调用save才会保存
        # print("i am write_row function")
        if isinstance(data_list_or_dic, list):
            self.book[sheet_name].append(data_list_or_dic)
            self.FirstRow = False
            self.count = self.count + 1

            # self.FirstRow = False
        elif isinstance(data_list_or_dic, pd.Series):
            self.book[sheet_name].append(data_list_or_dic.to_list())
            self.FirstRow = False
            self.count = self.count + 1
        else:
            self.unpack_dic(data_list_or_dic, sheet_name)
            # print(self.book[sheet_name].max_row)
            self.count = self.count + 1

    def save(self):
        # 基于openxyl保存后无需重新载入
        self.book.save(self.filepath_and_name)

    def run(self):
        sheet = self.Work_on_sheet
        save = self.Save_in_Write
        save_row = self.Save_each_rows
        if not self.Write_many_sheets:
            while 1:
                # sheet_name =
                # print("i am alive run excel")
                if save:
                    # print("i am alive")
                    # print("get"*15,self.queue_writer.get())
                    self.write_row_Save(data_list_or_dic=self.queue_writer.get(), sheet_name=sheet, save_row=save_row)
                    if not self.control.empty():
                        self.save()
                        self.control.get()
                        self.terminate()
                else:
                    # print("ima in no save")
                    self.write_row(sheet_name=sheet, data_list_or_dic=self.queue_writer.get())
                    if not self.control.empty():
                        self.save()
                        self.control.get()
                        self.terminate()

    def unpack_dic(self, data_dic, sheet_name):
        if self.FirstRow:
            temp_ItemIndex = []
            temp_ItemValue = []
            for k, v in data_dic.items():
                temp_ItemIndex.append(k)
                temp_ItemValue.append(v)
                # if temp_ItemIndex.__len__() > self.ItemIndex.__len__():
                #     diff = set(temp_ItemIndex)&
            self.ItemIndex[sheet_name] = temp_ItemIndex
            self.book[sheet_name].append(temp_ItemIndex)
            self.FirstRow = False
            self.book[sheet_name].append(temp_ItemValue)
        else:
            self.book[sheet_name].append(list(data_dic.values()))

    '''
    panda读写，由于panda读写保存后需要重新调用self.writer = pd.ExcelWriter(self.filepath_and_name)开辟文件句柄
    故具有劣势，这里暂时不写成线程
    '''

    def panda_write(self, panda_data, sheet_name="Sheet1"):
        # 必须调用 panda_flush 才可以实现保存修改
        self.writer = pd.ExcelWriter(self.filepath_and_name)
        panda_data.to_excel(self.writer, startrow=self.sheet_row_cloum[sheet_name] + 1)

    def panda_flush(self):
        self.writer.save()
        self.writer = pd.ExcelWriter(self.filepath_and_name, engine="openpyxl")
        self.sheet_row_cloum = {str(temp): self.book[temp].max for temp in self.sheets_names}

    def read(self, sheet_name=None, stop_row=None, start_row=0):
        """
        读取一个sheet表，从start_row到Stop_row 不指定stop则readall 不指定start则从第一个有数据的位置读
        \n返回嵌套的列表 [ [..]， [..] ]
        """
        sheet_name = self.book.sheetnames[0] if sheet_name is None else sheet_name
        sheet = self.book[sheet_name]
        index, min_row, max_row, min_cloumn, max_cloumn = self.read_title(sheet)
        start_row = min_row + 1 if start_row <= min_row + 1 else start_row
        stop_row = max_row if stop_row is None else stop_row
        sheet_slice = sheet[
                      c.get_column_letter(min_cloumn) + str(start_row):c.get_column_letter(max_cloumn) + str(stop_row)]
        data = [index]
        for row in sheet_slice:
            data.append(self.unpack_one_row(row))
        return data

    def read_with_index(self, *myindex, sheet_name=None, start_row=None, stop_row=None):
        """
        读取部分index对应的列
        :param index:
        :param sheet_name:
        :param start_row:
        :param stop_row:
        :return: 返回嵌套列表[ [..] , [..] ]
        """

        sheet_name = self.book.sheetnames[0] if sheet_name is None else sheet_name
        sheet = self.book[sheet_name]
        index, min_row, max_row, min_cloumn, max_cloumn = self.read_title(sheet)
        data = []
        start_row = start_row if start_row is not None else min_row
        stop_row = stop_row if stop_row is not None else max_row
        for x in range(stop_row - start_row):
            row = x + start_row
            row_str = str(row)
            data_row = []
            for one_index in myindex:
                clo = index.index(one_index) + min_cloumn
                clo_letter = c.get_column_letter(clo)
                data_row.append(sheet[clo_letter + row_str].value)
            data.append(data_row)
        return data

    def getStartRowAndStopRow(self,sheet_name=None, start_row=None, stop_row=None):
        sheet_name = self.book.sheetnames[0] if sheet_name is None else sheet_name
        sheet = self.book[sheet_name]
        index, min_row, max_row, min_cloumn, max_cloumn = self.read_title(sheet)
        start_row = start_row if start_row is not None else min_row
        stop_row = stop_row if stop_row is not None else max_row
        return start_row,stop_row, index, min_row, max_row, min_cloumn, max_cloumn

    def operate(self, *myindex, savePath=None, f= None,sheet_name=None, start_row=None, stop_row=None):
        sheet_name = self.sheets_names[0] if sheet_name is None else sheet_name
        start_row, stop_row ,index, min_row, max_row, min_cloumn, max_cloumn = self.getStartRowAndStopRow(sheet_name,start_row,stop_row)
        for i in myindex:
            for rowsNum in range(stop_row - start_row):
                rowsNum = start_row + rowsNum + 1
                colum_number = min_cloumn+index.index(i)
                cell_number = c.get_column_letter(colum_number)+str(rowsNum)
                if f is None:
                    self.book[sheet_name][cell_number] = self.modify(
                        self.book[sheet_name][cell_number].value,rowsNum,colum_number)
                else:
                    self.book[sheet_name][cell_number] = f(
                        self.book[sheet_name][cell_number].value, rowsNum, colum_number)
        self.book.save(savePath) if savePath else self.book.save(self.filepath_and_name)


    def modify(self,value:str,row,colum):

        return value




    def unpack_one_row(self, row):
        temp_list = []
        for x in row:
            # print(x)
            temp_list.append(x.value)
        return temp_list

    def read_title(self, sheet):
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_cloumn = sheet.min_column
        max_cloumn = sheet.max_column
        clo_start_letter = c.get_column_letter(min_cloumn)
        clo_end_letter = c.get_column_letter(max_cloumn)

        index_c = sheet[clo_start_letter + str(min_row):clo_end_letter + str(min_row)]
        index = [x.value for x in index_c[0]]
        return index, min_row, max_row, min_cloumn, max_cloumn


if __name__ == "__main__":
    e = Excel("D:\\Users\\wy\\Desktop\\demo.xlsx")
    e.operate("a",f=lambda v,r,c:(int(v)*int(v)))
